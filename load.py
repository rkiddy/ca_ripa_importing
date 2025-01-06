
from openpyxl import load_workbook

import argparse

from sqlalchemy import create_engine

import config

cfg = config.cfg()

engine = create_engine(f"mysql+pymysql://{cfg['USR']}:{cfg['PWD']}@{cfg['HOST']}/{cfg['DB']}")
conn = engine.connect()


def db_exec(engine, sql):
    # print(f"sql: {sql}")
    if sql.strip().startswith('select'):
        return [dict(r) for r in engine.execute(sql).fetchall()]
    else:
        return engine.execute(sql)


def open_workbook(path):
    workbook = load_workbook(filename=path)
    print(f"Worksheet names: {workbook.sheetnames}")
    sheet = workbook.active
    print(sheet)
    print(f"The title of the Worksheet is: {sheet.title}")
    print(f'The value of A3 is {sheet["A3"].value}')
    print(f'The value of D3 is {sheet["D3"].value}')
    print(f'The value of F3 is {sheet["F3"].value}')
    cell = sheet['H3']
    print(f'The variable "cell" is {cell.value}')


alphas = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
alpha2 = "ABCDEFGHIJKL"

columns = list(alphas)
columns.extend([f"A{r}" for r in list(alphas)])
columns.extend([f"B{r}" for r in list(alphas)])
columns.extend([f"C{r}" for r in list(alphas)])
columns.extend([f"D{r}" for r in list(alphas)])
columns.extend([f"E{r}" for r in list(alpha2)])

# print(f"columns: {columns}")

char_cols = ['A', 'C', 'D', 'E', 'H', 'I', 'J', 'EI', 'EJ', 'EK', 'EL']

if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('--year')
    args = parser.parse_args()

    if not args.year:
        sql = "select * from ripa_files"
    else:
        sql = f"select * from ripa_files where path like '%% {args.year} %%'"
    files = db_exec(conn, sql)
    # files = [{'pk': 0, 'path': 'test.xlsx'}]

    pk = db_exec(conn, f"select max(pk) as pk from ripa_{args.year}")[0]['pk']
    if pk is None:
        pk = 0

    for r in files:
        path = r['path']
        fpk = int(r['pk'])
        print(f"path: {path}")

        sql = f"select count(0) as count from ripa_{args.year} where file_pk = {fpk}"
        count = db_exec(conn, sql)[0]['count']

        if count > 0:
            print(f"found rows # {count}. skipping.")
            continue

        workbook = load_workbook(filename=path, read_only=True)
        sheet = workbook.active

        done = False
        first_row = True
        row_num = 0
        sqls = list()

        for row in sheet.iter_rows():

            data = list()

            if first_row:
                first_row = False
                continue

            for cell in row:
                data.append(f"{cell.value}")

            cdata = dict(zip(columns, data))

            for char_col in char_cols:
                value = cdata[char_col].replace("'", "''")
                cdata[char_col] = f"'{value}'"

            for col in columns:
                if cdata[col] == 'None' or cdata[col] == "'None'":
                    cdata[col] = 'NULL'

            cdata['F'] = f"'{str(cdata['F'])[:10]}'"

            cdata = list(cdata.values())

            pk += 1
            row_num += 1

            cdata.insert(0, f"{fpk}")
            cdata.insert(0, f"{pk}")

            sqls.append(f"({','.join(cdata)})")

            if len(sqls) > 100:
                values = ','.join(sqls)
                sql = f"insert into ripa_{args.year} values {values}"
                db_exec(conn, sql)
                sqls = list()

            if (row_num % 10000) == 0:
                print(f"{row_num}...")

        workbook.close()

        if len(sqls) > 100:
            values = ','.join(sqls)
            sql = f"insert into ripa_{args.year} values {values}"
            db_exec(conn, sql)
            sqls = list()
